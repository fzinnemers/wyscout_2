import pandas as pd
import numpy as np
import openpyxl
import re

from unidecode import unidecode
from datetime import datetime
from openpyxl.styles import Font, PatternFill

from wyscout_etl.calculate_totals import CalculateKPI


class ScoutingExcel:

    def __init__(self):
        pass

    def _create_scouting_excel(
        self,
        df, 
        sink_path, 
        kpi_method = "general.py",
        general_variables=[
            "full_name",
            "birth_date",
            "birth_country_name",
            "foot",
            "passport_country_names1",
            "height",
            "last_club_name",
            "division",
            "league_country",
            "league_competition",
            "total_matches",
            "minutes_on_field",
            "main_position",
            "primary_position",
        ],
    ):
        
        
        df = df[df["minutes_on_field"] > 46]

        # Clean name and age 
        df["full_name"] = df["full_name"].apply(self._clean_name)
        df["birth_date"] = df["birth_date"].apply(self._calculate_age)


        # Get settings from 
        get_settings = CalculateKPI()

        # Get KPI scoring values from db settings 
        kpi_scoring_values = get_settings._import_variables_from_script(kpi_method)[1]

        zscore_df, zscore_df_padj, quantile_df, quantile_df_padj = (
            self._create_dataframes(
                df=df,
                general_variables=general_variables,
                kpi_scoring_values=kpi_scoring_values,
            )
        )


        zscore_df = self.rename_score_columns(zscore_df, kpi_scoring_values)
        zscore_df_padj = self.rename_score_columns(zscore_df_padj, kpi_scoring_values)
        quantile_df = self.rename_score_columns(quantile_df, kpi_scoring_values)
        quantile_df_padj = self.rename_score_columns(quantile_df_padj, kpi_scoring_values)

        zscore_df = zscore_df.reset_index(drop = True)
        zscore_df_padj = zscore_df_padj.reset_index(drop = True)
        quantile_df = quantile_df.reset_index(drop = True)
        quantile_df_padj = quantile_df_padj.reset_index(drop = True)


        print("Begin writing to Excel")
        self._write_to_excel(
            dataframes=[
                zscore_df,
                zscore_df_padj,
                quantile_df,
                quantile_df_padj,
            ],
            sink_path=sink_path,
            general_variables=general_variables
        )

    def _write_to_excel(
        self,
        dataframes,
        sink_path: str,
        general_variables,
        sheetnames=None,
    ):
        if sheetnames is None:
            sheetnames = ["zscore", "zscore_padj", "quantile", "quantile_padj"]

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()

        for i, df in enumerate(dataframes):
            sheet = workbook.create_sheet(title=sheetnames[i])

            threshold_columns = [i for i in df.columns if i not in general_variables]
            threshold_df = self._create_threshold_dfs(df, score_columns = threshold_columns)

            # Add column names as the first row
            for j, column_name in enumerate(df.columns, start=1):
                cell = sheet.cell(row=1, column=j, value=column_name)
                # Apply bold font to the first row
                cell.font = Font(bold=True)

            # Writing away the numbers
            for j, column_name in enumerate(df.columns, start=1):
                if column_name in threshold_columns:
                    temp_threshold_df = threshold_df[
                        threshold_df["column"] == column_name
                    ]

                    for k, value in enumerate(df[column_name], start=2):
                        cell = sheet.cell(row=k, column=j, value=value)
                        position = df["main_position"][k - 2]
                        temp = temp_threshold_df[
                            temp_threshold_df["position"] == position
                        ]

                        high_threshold_value = temp.high_threshold.values[0]
                        low_threshold_value = temp.low_threshold.values[0]

                        # Apply formatting if the value is above/below the threshold
                        if isinstance(value, (int, float)) and not np.isnan(value):
                            if value > high_threshold_value:
                                cell.fill = PatternFill(
                                    start_color="00FF00",
                                    end_color="00FF00",
                                    fill_type="solid",
                                )
                            elif value < low_threshold_value:
                                cell.fill = PatternFill(
                                    start_color="FF0000",
                                    end_color="FF0000",
                                    fill_type="solid",
                                )
                        else:
                            cell.fill = PatternFill(
                                start_color="FFFFFF",
                                end_color="FFFFFF",
                                fill_type="solid",
                            )
                else:
                    # If the column is not in threshold_columns, directly copy values
                    for k, value in enumerate(df[column_name], start=2):
                        sheet.cell(row=k, column=j, value=value)

            # Freeze the first row
            sheet.freeze_panes = sheet["A2"]

        # Remove the default sheet created with the workbook
        workbook.remove(workbook["Sheet"])

        # Save the Excel file
        workbook.save(sink_path)

   
    def _clean_name(self, name):
        # Remove accents
        cleaned_name = unidecode(name)

        # Remove non-alphanumeric characters except spaces
        cleaned_name = re.sub(r"[^a-zA-Z0-9\s]", "", cleaned_name)

        # Strip double spaces
        cleaned_name = re.sub(r"\s+", " ", cleaned_name)

        # Strip leading and trailing spaces
        cleaned_name = cleaned_name.strip()

        return cleaned_name

    def _calculate_age(self, birth_date):
        if isinstance(birth_date, str):
            try:
                birth_date = datetime.strptime(birth_date, "%Y-%m-%d")
                today = datetime.now()
                age_days = (today - birth_date).days
                age_years = age_days / 365.25  # Taking leap years into account
                age = round(age_years, 2)
                return age
            except ValueError:
                # Handle invalid date format
                return None
        else:
            # Handle non-string values, such as NaN
            return None

    def _create_threshold_dfs(self, df, score_columns):
        data = []
        for position in set(df["main_position"]):
            for column in score_columns:
                high_threshold = np.nanquantile(
                    df[df["main_position"] == position][column], 0.8
                )
                low_threshold = np.nanquantile(
                    df[df["main_position"] == position][column], 0.2
                )

                data.append([position, column, high_threshold, low_threshold])

        threshold_df = pd.DataFrame(
            data, columns=["position", "column", "high_threshold", "low_threshold"]
        )

        return threshold_df


    def _create_dataframes(self, df, kpi_scoring_values, general_variables):

        # Get all the KPI defined in the scoring value file
        all_kpi = list(kpi_scoring_values.keys())

        # Get all the variables that define the KPI
        variables_of_kpi = []

        for kpi, variables in kpi_scoring_values.items():
            for variable, weight in variables.items():
                variables_of_kpi.append(f"{variable}")

        # add the specific strings to the kpi names
        zscore_kpi_columns = ["avg_zscore_" + i for i in all_kpi]
        zscore_kpi_padj_columns = ["avg_zscore_" + i + "_padj" for i in all_kpi]

        quantile_kpi_columns = ["avg_quantile_" + i for i in all_kpi]
        quantile_kpi_padj_columns = ["avg_quantile_" + i + "_padj" for i in all_kpi]

        # add the specific strings to the kpi variable names
        zscore_kpi_variables = ["zscore_" + i for i in variables_of_kpi]

        zscore_kpi_padj_variables = [
            (
                "zscore_" + i + "_padj"
                if "zscore_" + i + "_padj" in df.columns
                else "zscore_" + i
            )
            for i in variables_of_kpi
        ]

        quantile_kpi_variables = ["quantile_" + i for i in variables_of_kpi]

        quantile_kpi_padj_variables = [
            (
                "quantile_" + i + "_padj"
                if "quantile_" + i + "_padj" in df.columns
                else "quantile_" + i
            )
            for i in variables_of_kpi
        ]

        # Final column order for the datasheets
        zscore_df = df[
            general_variables
            + zscore_kpi_columns
            + ["weighted_zscore_total"]
            + zscore_kpi_variables
        ]
        zscore_df_padj = df[
            general_variables
            + zscore_kpi_padj_columns
            + ["weighted_zscore_total_padj"]
            + zscore_kpi_padj_variables
        ]

        quantile_df = df[
            general_variables
            + quantile_kpi_columns
            + ["weighted_quantile_total"]
            + quantile_kpi_variables
        ]
        quantile_df_padj = df[
            general_variables
            + quantile_kpi_padj_columns
            + ["weighted_quantile_total_padj"]
            + quantile_kpi_padj_variables
        ]

        return zscore_df, zscore_df_padj, quantile_df, quantile_df_padj
    

    def rename_score_columns(self, df, kpi_scoring_dict): 

        # All values that need replacing to make the score columns more readable 
        to_replace_values = ['avg_zscore_', 'avg_quantile_','_padj', 'weighted_zscore_', 'weighted_quantile_', 'zscore_', 'quantile_']
        for i in to_replace_values:
            df.columns = [column.replace(i, '') for column in df.columns]


        # Extract from the kpi_scoring_dict the name of the variables and clean that name
        new_variables = []
        count = 1
        for main_key, sub_dict in kpi_scoring_dict.items(): 
            for sub_key, weight in sub_dict.items():
                sub_variable = f"{sub_key} - weight: {weight} - KPI: {count}"
                new_variables.append(sub_variable.replace('_avg', ''))
            count += 1

        # Rename the columns 
        new_columns = list(df.columns[:len(df.columns) - len(new_variables)]) + new_variables
        df.columns = new_columns
        
        return df
